#!/usr/bin/env python3
"""
Excel Formatter Script

A simple yet robust Python script that processes Excel files from input_files directory
using template files and applies mapping rules to generate formatted output files.

Author: Assistant
Date: 2025
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import xlrd
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
import logging
from datetime import datetime
import sys
import os
import json

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('excel_formatter.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


class ExcelFormatter:
    """Main class for Excel file formatting operations."""
    
    def __init__(self, input_dir: str = "../input_files", 
                 template_dir: str = "../template_files", 
                 output_dir: str = "../output_files"):
        """
        Initialize the Excel formatter.
        
        Args:
            input_dir: Directory containing input Excel files
            template_dir: Directory containing template Excel files
            output_dir: Directory for output files
        """
        self.input_dir = Path(input_dir)
        self.template_dir = Path(template_dir)
        self.output_dir = Path(output_dir)
        
        # Ensure directories exist
        self.input_dir.mkdir(exist_ok=True)
        self.template_dir.mkdir(exist_ok=True)
        self.output_dir.mkdir(exist_ok=True)
        
        # Load configuration from external file
        self.config = self.load_configuration()
        
        logger.info("ExcelFormatter initialized successfully")
    
    def load_configuration(self) -> Dict[str, Any]:
        """Load configuration from JSON file."""
        try:
            config_path = Path(__file__).parent / "mapping_config.json"
            with open(config_path, 'r') as f:
                config = json.load(f)
            logger.info("Configuration loaded successfully")
            return config
        except Exception as e:
            logger.error(f"Error loading configuration: {str(e)}")
            # Return default configuration
            return self.get_default_config()
    
    def get_default_config(self) -> Dict[str, Any]:
        """Get default configuration if file loading fails."""
        return {
            "output_columns": [
                {"name": "Employee Name", "source_column": "Name", "alignment": "left"},
                {"name": "Check #", "source_column": "", "alignment": "center"},
                {"name": "Chk Amt", "source_column": "Net pay", "alignment": "right"},
                {"name": "Gross", "source_column": "Adjusted gross", "alignment": "right"},
                {"name": "Fica E/R", "source_column": "Employee taxes - SS + Employee taxes - Med", "alignment": "right"},
                {"name": "Liab", "source_column": "=Chk Amt - Gross - Fica", "alignment": "right"},
                {"name": "Date", "source_column": "Pay date", "alignment": "center"},
                {"name": "Period", "source_column": "Time period", "alignment": "center"}
            ]
        }
    
    def read_excel_file(self, file_path: Path) -> pd.DataFrame:
        """
        Read Excel file and return DataFrame.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            pandas DataFrame containing the data
            
        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file format is not supported
        """
        try:
            if not file_path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")
            
            logger.info(f"Reading file: {file_path}")
            
            if file_path.suffix.lower() == '.xls':
                # Try to find the header row by looking for common payroll column names
                df_raw = pd.read_excel(file_path, engine='xlrd', header=None)
                
                # Look for a row that contains common payroll headers
                header_row = None
                for idx, row in df_raw.iterrows():
                    row_str = ' '.join([str(cell) for cell in row if pd.notna(cell)]).lower()
                    # Look for more specific payroll terms
                    if any(keyword in row_str for keyword in ['employee name', 'net pay', 'gross pay', 'pay date', 'time period', 'social security', 'medicare']):
                        header_row = idx
                        logger.info(f"Found potential header row {idx} with content: {row_str[:100]}")
                        break
                    # Also check for individual cells that might be headers
                    elif any(str(cell).lower() in ['name', 'employee', 'pay', 'gross', 'net', 'date', 'period'] for cell in row if pd.notna(cell)):
                        header_row = idx
                        logger.info(f"Found potential header row {idx} with individual headers")
                        break
                
                if header_row is not None:
                    df = pd.read_excel(file_path, engine='xlrd', header=header_row)
                    logger.info(f"Found headers at row {header_row}")
                else:
                    df = df_raw
                    logger.warning("Could not find header row, using raw data")
                    
            elif file_path.suffix.lower() in ['.xlsx', '.xlsm']:
                df = pd.read_excel(file_path, engine='openpyxl')
            else:
                raise ValueError(f"Unsupported file format: {file_path.suffix}")
            
            logger.info(f"Successfully read {len(df)} rows from {file_path.name}")
            return df
            
        except Exception as e:
            logger.error(f"Error reading file {file_path}: {str(e)}")
            raise
    
    def analyze_template(self, template_path: Path) -> Dict[str, Any]:
        """
        Analyze template file to understand its structure.
        
        Args:
            template_path: Path to the template file
            
        Returns:
            Dictionary containing template information
        """
        try:
            logger.info(f"Analyzing template: {template_path}")
            
            wb = load_workbook(template_path, data_only=False)
            ws = wb.active
            
            # Get column headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value))
                else:
                    headers.append("")
            
            # Find formulas in the sheet
            formulas = {}
            for row in ws.iter_rows(min_row=2, max_row=min(10, ws.max_row)):
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        col_letter = cell.column_letter
                        if col_letter not in formulas:
                            formulas[col_letter] = cell.value
            
            template_info = {
                "headers": headers,
                "formulas": formulas,
                "sheet_name": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column
            }
            
            logger.info(f"Template analysis complete. Headers: {len(headers)}")
            return template_info
            
        except Exception as e:
            logger.error(f"Error analyzing template {template_path}: {str(e)}")
            raise
    
    def apply_mapping(self, input_df: pd.DataFrame, template_info: Dict[str, Any]) -> pd.DataFrame:
        """
        Apply mapping rules to transform input data according to configuration.
        
        Args:
            input_df: Input DataFrame
            template_info: Template structure information
            
        Returns:
            Transformed DataFrame matching configuration structure
        """
        try:
            logger.info("Applying mapping rules to input data")
            
            # Create output DataFrame based on configuration
            output_data = {}
            output_columns = self.config["output_columns"]
            
            # Apply mapping rules from configuration
            for col_config in output_columns:
                col_name = col_config["name"]
                source_col = col_config["source_column"]
                
                if source_col.startswith("="):
                    # Handle formula
                    output_data[col_name] = self._evaluate_formula(
                        source_col, input_df, output_data
                    )
                elif source_col == "":
                    # Blank field
                    output_data[col_name] = [""] * len(input_df)
                else:
                    # Direct column mapping
                    if source_col in input_df.columns:
                        output_data[col_name] = input_df[source_col].values
                    else:
                        logger.warning(f"Column '{source_col}' not found in input data")
                        output_data[col_name] = [""] * len(input_df)
            
            result_df = pd.DataFrame(output_data)
            
            # Apply void filtering if enabled
            result_df = self.apply_void_filtering(result_df)
            
            logger.info(f"Mapping applied successfully. Output shape: {result_df.shape}")
            return result_df
            
        except Exception as e:
            logger.error(f"Error applying mapping: {str(e)}")
            raise
    
    def apply_void_filtering(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Filter out rows where specified columns are all zero.
        
        Args:
            df: DataFrame to filter
            
        Returns:
            Filtered DataFrame
        """
        try:
            void_config = self.config.get("void", {})
            
            if not void_config.get("enabled", False):
                logger.info("Void filtering is disabled")
                return df
            
            zero_columns = void_config.get("zero_columns", [])
            if not zero_columns:
                logger.info("No zero columns specified for void filtering")
                return df
            
            # Check which columns exist in the DataFrame
            existing_columns = [col for col in zero_columns if col in df.columns]
            if not existing_columns:
                logger.warning(f"None of the specified zero columns {zero_columns} found in DataFrame")
                return df
            
            logger.info(f"Applying void filtering on columns: {existing_columns}")
            
            # Create a mask for rows where all specified columns are zero or empty
            mask = pd.Series([True] * len(df))
            
            for col in existing_columns:
                # Convert to numeric, replacing non-numeric values with 0
                numeric_col = pd.to_numeric(df[col], errors='coerce').fillna(0)
                # Check if column is zero
                mask = mask & (numeric_col == 0)
            
            # Count rows to be removed
            rows_to_remove = mask.sum()
            total_rows = len(df)
            
            if rows_to_remove > 0:
                logger.info(f"Removing {rows_to_remove} void rows out of {total_rows} total rows")
                # Keep rows that are NOT all zero
                filtered_df = df[~mask].copy()
                logger.info(f"Filtered DataFrame shape: {filtered_df.shape}")
                return filtered_df
            else:
                logger.info("No void rows found to remove")
                return df
                
        except Exception as e:
            logger.error(f"Error applying void filtering: {str(e)}")
            return df
    
    def _evaluate_formula(self, formula: str, input_df: pd.DataFrame, 
                         output_data: Dict[str, List]) -> List[Any]:
        """
        Evaluate Excel formula with input data.
        
        Args:
            formula: Formula string (e.g., "=Chk Amt - Gross - Fica")
            input_df: Input DataFrame
            output_data: Current output data for reference
            
        Returns:
            List of calculated values
        """
        try:
            result = []
            formula_clean = formula.replace("=", "").strip()
            
            for idx, row in input_df.iterrows():
                try:
                    # Replace column references with actual values
                    formula_eval = formula_clean
                    
                    # Map formula column names to actual data
                    column_mapping = {
                        "Chk Amt": "Net pay",
                        "Gross": "Adjusted gross", 
                        "Fica": "Employee taxes - SS + Employee taxes - Med"
                    }
                    
                    for formula_col, actual_col in column_mapping.items():
                        if actual_col in input_df.columns:
                            value = row[actual_col]
                            # Handle non-numeric values
                            try:
                                numeric_value = float(value) if pd.notna(value) else 0
                                formula_eval = formula_eval.replace(formula_col, str(numeric_value))
                            except (ValueError, TypeError):
                                formula_eval = formula_eval.replace(formula_col, "0")
                        else:
                            formula_eval = formula_eval.replace(formula_col, "0")
                    
                    # Evaluate the formula safely
                    if any(op in formula_eval for op in ["+", "-", "*", "/"]):
                        calculated_value = eval(formula_eval)
                        result.append(calculated_value)
                    else:
                        result.append(0)
                        
                except Exception as e:
                    logger.warning(f"Error evaluating formula for row {idx}: {e}")
                    result.append(0)
            
            return result
            
        except Exception as e:
            logger.error(f"Error evaluating formula {formula}: {str(e)}")
            return [0] * len(input_df)
    
    def save_output(self, df: pd.DataFrame, output_path: Path) -> None:
        """
        Save DataFrame to Excel file with formatting.
        
        Args:
            df: DataFrame to save
            output_path: Output file path
        """
        try:
            logger.info(f"Saving output to: {output_path}")
            
            # Save to Excel first
            df.to_excel(output_path, index=False, engine='openpyxl')
            
            # Apply formatting
            self.apply_formatting(output_path, df)
            
            logger.info(f"Output saved successfully: {output_path.name}")
            
        except Exception as e:
            logger.error(f"Error saving output file {output_path}: {str(e)}")
            raise
    
    def apply_formatting(self, output_path: Path, df: pd.DataFrame) -> None:
        """
        Apply formatting to the Excel file based on configuration.
        
        Args:
            output_path: Path to the Excel file
            df: DataFrame that was saved
        """
        try:
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Apply header formatting
            self.format_headers(ws, df)
            
            # Apply column formatting
            self.format_columns(ws, df)
            
            # Apply general settings
            self.apply_general_settings(ws)
            
            wb.save(output_path)
            logger.info("Formatting applied successfully")
            
        except Exception as e:
            logger.error(f"Error applying formatting: {str(e)}")
            raise
    
    def format_headers(self, ws, df: pd.DataFrame) -> None:
        """Format the header row."""
        try:
            header_config = self.config.get("header_formatting", {})
            column_name_alignment = self.config.get("column_name_alignment", {})
            
            # Header row styling
            header_font = Font(
                bold=header_config.get("bold", True),
                color=header_config.get("font_color", "FFFFFF")
            )
            
            header_fill = PatternFill(
                start_color=header_config.get("background_color", "366092"),
                end_color=header_config.get("background_color", "366092"),
                fill_type="solid"
            )
            
            # Default header alignment
            default_header_alignment = Alignment(
                horizontal=header_config.get("alignment", "center"),
                vertical="center"
            )
            
            # Apply to header row
            for col_num in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                
                # Use specific column alignment if defined, otherwise use default
                col_name = df.columns[col_num - 1]
                if col_name in column_name_alignment:
                    cell.alignment = Alignment(
                        horizontal=column_name_alignment[col_name],
                        vertical="center"
                    )
                else:
                    cell.alignment = default_header_alignment
                
        except Exception as e:
            logger.warning(f"Error formatting headers: {str(e)}")
    
    def format_columns(self, ws, df: pd.DataFrame) -> None:
        """Format data columns based on configuration."""
        try:
            output_columns = self.config["output_columns"]
            
            for col_idx, col_config in enumerate(output_columns, 1):
                col_name = col_config["name"]
                formatting = col_config.get("formatting", {})
                
                # Set column width
                if "width" in col_config:
                    ws.column_dimensions[get_column_letter(col_idx)].width = col_config["width"]
                
                # Set alignment
                alignment = Alignment(
                    horizontal=col_config.get("alignment", "left"),
                    vertical="center"
                )
                
                # Apply formatting to all data rows
                for row_num in range(2, len(df) + 2):  # Start from row 2 (after header)
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.alignment = alignment
                    
                    # Apply number formatting
                    if "number_format" in formatting:
                        cell.number_format = formatting["number_format"]
                    
                    # Apply negative number formatting
                    if "negative_format" in formatting and isinstance(cell.value, (int, float)) and cell.value < 0:
                        cell.number_format = formatting["negative_format"]
                    
                    # Apply date formatting
                    if "date_format" in formatting and col_name in ["Date"]:
                        cell.number_format = "mm/dd/yyyy"
                    
                    # Apply date range formatting for Period column
                    if "date_range_format" in formatting and col_name in ["Period"]:
                        # This would need special handling for date ranges
                        pass
                        
        except Exception as e:
            logger.warning(f"Error formatting columns: {str(e)}")
    
    def apply_general_settings(self, ws) -> None:
        """Apply general worksheet settings."""
        try:
            general_settings = self.config.get("general_settings", {})
            
            # Auto-fit columns
            if general_settings.get("auto_fit_columns", True):
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze panes
            if "freeze_panes" in general_settings:
                ws.freeze_panes = general_settings["freeze_panes"]
                
        except Exception as e:
            logger.warning(f"Error applying general settings: {str(e)}")
    
    def process_single_file(self, input_file: Path, template_file: Path) -> Path:
        """
        Process a single input file with a template.
        
        Args:
            input_file: Path to input file
            template_file: Path to template file
            
        Returns:
            Path to the generated output file
        """
        try:
            logger.info(f"Processing file: {input_file.name}")
            
            # Read input data
            input_df = self.read_excel_file(input_file)
            
            # Print available columns for debugging
            logger.info(f"Available columns in {input_file.name}: {input_df.columns.tolist()}")
            logger.info(f"Total columns: {len(input_df.columns)}")
            logger.info(f"First few rows of data:\n{input_df.head()}")
            
            # Analyze template
            template_info = self.analyze_template(template_file)
            
            # Apply mapping
            output_df = self.apply_mapping(input_df, template_info)
            
            # Generate output filename with .xlsx extension and timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"{input_file.stem}_formatted_{timestamp}.xlsx"
            output_path = self.output_dir / output_filename
            
            # Save output
            self.save_output(output_df, output_path)
            
            logger.info(f"Successfully processed: {input_file.name} -> {output_filename}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error processing file {input_file.name}: {str(e)}")
            raise
    
    def process_all_files(self) -> List[Path]:
        """
        Process all input files with available templates.
        
        Returns:
            List of generated output file paths
        """
        try:
            logger.info("Starting batch processing of all files")
            
            # Get all input files
            input_files = list(self.input_dir.glob("*.xls*"))
            if not input_files:
                logger.warning("No input files found")
                return []
            
            # Get template files
            template_files = list(self.template_dir.glob("*.xls*"))
            if not template_files:
                logger.warning("No template files found")
                return []
            
            # Use the first template for all input files
            template_file = template_files[0]
            logger.info(f"Using template: {template_file.name}")
            
            output_files = []
            
            for input_file in input_files:
                try:
                    output_path = self.process_single_file(input_file, template_file)
                    output_files.append(output_path)
                except Exception as e:
                    logger.error(f"Failed to process {input_file.name}: {str(e)}")
                    continue
            
            logger.info(f"Batch processing complete. Generated {len(output_files)} output files")
            return output_files
            
        except Exception as e:
            logger.error(f"Error in batch processing: {str(e)}")
            raise


def main():
    """Main function to run the Excel formatter."""
    try:
        logger.info("Starting Excel Formatter")
        
        # Initialize formatter
        formatter = ExcelFormatter()
        
        # Process all files
        output_files = formatter.process_all_files()
        
        if output_files:
            print(f"\n[SUCCESS] Successfully processed {len(output_files)} files:")
            for output_file in output_files:
                print(f"   - {output_file.name}")
        else:
            print("\n[ERROR] No files were processed")
        
        logger.info("Excel Formatter completed successfully")
        
    except Exception as e:
        logger.error(f"Excel Formatter failed: {str(e)}")
        print(f"\n[ERROR] {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
