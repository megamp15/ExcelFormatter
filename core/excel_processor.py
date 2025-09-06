#!/usr/bin/env python3
"""
Core Excel processing logic for Excel Formatter application.

This module contains the core logic for reading, processing, and formatting
Excel files based on configuration mappings.
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import xlrd
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
import logging
from datetime import datetime
import re

from config.settings import *


class ExcelProcessor:
    """Core Excel file processing class."""
    
    def __init__(self):
        """Initialize the Excel processor."""
        self.logger = logging.getLogger(__name__)
        
    def read_excel_file(self, file_path: Path) -> pd.DataFrame:
        """
        Read Excel file and return DataFrame.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            pandas DataFrame containing the data
            
        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file format is not supported
        """
        try:
            if not file_path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")
            
            self.logger.info(f"Reading file: {file_path}")
            
            if file_path.suffix.lower() == '.xls':
                # Handle legacy XLS files - try to find header row
                df_raw = pd.read_excel(file_path, engine='xlrd', header=None)
                
                # Look for header row with common Excel column patterns
                header_row = self._find_header_row(df_raw)
                
                if header_row is not None:
                    df = pd.read_excel(file_path, engine='xlrd', header=header_row)
                    self.logger.info(f"Found headers at row {header_row}")
                else:
                    # Use first row as header if no patterns found
                    df = pd.read_excel(file_path, engine='xlrd', header=0)
                    self.logger.warning("Using first row as header")
                    
            elif file_path.suffix.lower() in ['.xlsx', '.xlsm']:
                df = pd.read_excel(file_path, engine='openpyxl')
            else:
                raise ValueError(f"Unsupported file format: {file_path.suffix}")
            
            # Clean column names
            df.columns = [str(col).strip() for col in df.columns]
            
            self.logger.info(f"Successfully read {len(df)} rows, {len(df.columns)} columns")
            return df
            
        except Exception as e:
            self.logger.error(f"Error reading file {file_path}: {str(e)}")
            raise
            
    def _find_header_row(self, df_raw: pd.DataFrame) -> Optional[int]:
        """
        Find the header row in a raw DataFrame by looking for common patterns.
        
        Args:
            df_raw: Raw DataFrame without headers
            
        Returns:
            Row index of headers, or None if not found
        """
        header_keywords = [
            'name', 'employee', 'pay', 'gross', 'net', 'date', 'period',
            'amount', 'salary', 'wage', 'hours', 'rate', 'deduction',
            'tax', 'social', 'security', 'medicare', 'federal', 'state'
        ]
        
        for idx, row in df_raw.iterrows():
            if idx > 20:  # Don't search too far down
                break
                
            # Convert row to string and check for keywords
            row_str = ' '.join([str(cell).lower() for cell in row if pd.notna(cell)])
            
            # Count matches
            matches = sum(1 for keyword in header_keywords if keyword in row_str)
            
            # If we find multiple matches, likely a header row
            if matches >= 3:
                return idx
                
        return None
        
    def apply_mapping(self, input_df: pd.DataFrame, config: Dict[str, Any]) -> pd.DataFrame:
        """
        Apply column mapping configuration to transform input data.
        
        Args:
            input_df: Input DataFrame
            config: Mapping configuration
            
        Returns:
            Transformed DataFrame
        """
        try:
            self.logger.info("Applying column mapping configuration")
            
            # Apply void filtering to input data first if enabled
            filtered_input_df = self._apply_void_filtering(input_df, config)
            
            output_data = {}
            output_columns = config.get("output_columns", [])
            
            for col_config in output_columns:
                col_name = col_config.get("name", "").strip()
                if not col_name:
                    continue
                    
                source_column = col_config.get("source_column", "")
                
                # Apply different mapping types
                if source_column.startswith("="):
                    # Formula mapping
                    output_data[col_name] = self._evaluate_formula(
                        source_column, filtered_input_df, output_data, col_config
                    )
                elif source_column == "":
                    # Blank column
                    output_data[col_name] = [""] * len(filtered_input_df)
                else:
                    # Direct column mapping
                    output_data[col_name] = self._map_column(
                        source_column, filtered_input_df, col_config
                    )
                    
            result_df = pd.DataFrame(output_data)
            
            # Apply column ordering if specified
            column_order = config.get("column_order", [])
            if column_order:
                # Filter to only include columns that exist in the result
                valid_order = [col for col in column_order if col in result_df.columns]
                # Add any remaining columns that weren't in the order
                remaining_cols = [col for col in result_df.columns if col not in valid_order]
                final_order = valid_order + remaining_cols
                result_df = result_df[final_order]
                self.logger.info(f"Applied column ordering: {final_order}")
            
            self.logger.info(f"Mapping applied successfully. Output shape: {result_df.shape}")
            return result_df
            
        except Exception as e:
            self.logger.error(f"Error applying mapping: {str(e)}")
            raise
            
    def _evaluate_formula(self, formula: str, input_df: pd.DataFrame, 
                         output_data: Dict[str, List], col_config: Dict[str, Any]) -> List[Any]:
        """
        Evaluate Excel-style formula.
        
        Args:
            formula: Formula string starting with '='
            input_df: Input DataFrame
            output_data: Current output data for reference
            col_config: Column configuration
            
        Returns:
            List of calculated values
        """
        try:
            result = []
            formula_clean = formula[1:].strip()  # Remove '='
            self.logger.info(f"Evaluating formula: {formula_clean}")
            
            for idx, row in input_df.iterrows():
                try:
                    # Replace column references with actual values
                    evaluated_formula = self._replace_formula_references(
                        formula_clean, row, output_data, idx
                    )
                    
                    # Debug logging for first few rows
                    if idx < 3:
                        self.logger.info(f"Row {idx}: Original: {formula_clean}, Evaluated: {evaluated_formula}")
                    
                    # Safely evaluate the formula
                    if self._is_safe_expression(evaluated_formula):
                        calculated_value = eval(evaluated_formula)
                        result.append(calculated_value)
                    else:
                        # Try to fix common issues with quotes
                        fixed_formula = evaluated_formula.replace('"', '')
                        if self._is_safe_expression(fixed_formula):
                            self.logger.info(f"Fixed expression for row {idx}: {fixed_formula}")
                            calculated_value = eval(fixed_formula)
                            result.append(calculated_value)
                        else:
                            self.logger.warning(f"Unsafe expression for row {idx}: {evaluated_formula}")
                            result.append(0)
                        
                except Exception as e:
                    self.logger.warning(f"Error evaluating formula for row {idx}: {e}")
                    result.append(0)
                    
            return result
            
        except Exception as e:
            self.logger.error(f"Error evaluating formula {formula}: {str(e)}")
            return [0] * len(input_df)
            
    def _replace_formula_references(self, formula: str, row: pd.Series, 
                                   output_data: Dict[str, List], row_idx: int) -> str:
        """
        Replace column references in formula with actual values.
        
        Args:
            formula: Formula string
            row: Current row data
            output_data: Current output data
            row_idx: Current row index
            
        Returns:
            Formula with replaced values
        """
        # Find all column references (words that could be column names)
        # This is a simplified approach - could be made more sophisticated
        import re
        
        # Look for column names - handle names with spaces and hyphens
        # First try to find quoted column names
        quoted_names = re.findall(r'"([^"]+)"', formula)
        # Then find unquoted column names (words with spaces, hyphens, etc.)
        unquoted_names = re.findall(r'\b[A-Za-z][A-Za-z0-9_\s\-]*[A-Za-z0-9_]\b', formula)
        
        # Combine and deduplicate
        words = list(set(quoted_names + unquoted_names))
        self.logger.info(f"Found words in formula: {words}")
        
        replaced_formula = formula
        
        for word in words:
            word_clean = word.strip()
            
            # Skip common operators and functions
            if word_clean.lower() in ['and', 'or', 'not', 'abs', 'sum', 'avg', 'max', 'min']:
                continue
                
            # Try to find matching column in input data first
            value = self._find_column_value(word_clean, row)
            
            # If not found in input data, try output data (previously calculated columns)
            if value is None and word_clean in output_data:
                try:
                    value = output_data[word_clean][row_idx]
                    self.logger.info(f"Found value for '{word_clean}' in output data: {value}")
                except (IndexError, KeyError):
                    value = None
            
            if value is not None:
                self.logger.info(f"Found value for '{word_clean}': {value}")
                # Replace the word with the numeric value
                # Use word boundaries for simple names, or exact match for complex names
                if ' ' in word_clean or '-' in word_clean:
                    # For names with spaces or hyphens, use exact match
                    replaced_formula = replaced_formula.replace(word_clean, str(value))
                else:
                    # For simple names, use word boundaries
                    replaced_formula = re.sub(
                        r'\b' + re.escape(word_clean) + r'\b', 
                        str(value), 
                        replaced_formula
                    )
            else:
                self.logger.warning(f"No value found for '{word_clean}'")
                
        return replaced_formula
        
    def _find_column_value(self, col_name: str, row: pd.Series) -> Optional[float]:
        """
        Find column value in row, with fuzzy matching.
        
        Args:
            col_name: Column name to find
            row: Row data
            
        Returns:
            Numeric value or None if not found
        """
        # Direct match
        if col_name in row.index:
            return self._to_numeric(row[col_name])
            
        # Case-insensitive match
        for col in row.index:
            if str(col).lower() == col_name.lower():
                return self._to_numeric(row[col])
                
        # Partial match
        for col in row.index:
            if col_name.lower() in str(col).lower() or str(col).lower() in col_name.lower():
                return self._to_numeric(row[col])
                
        return None
        
    def _to_numeric(self, value) -> float:
        """Convert value to numeric, returning 0 if conversion fails."""
        try:
            if pd.isna(value):
                return 0.0
            return float(value)
        except (ValueError, TypeError):
            return 0.0
            
    def _is_safe_expression(self, expr: str) -> bool:
        """
        Check if expression is safe to evaluate.
        
        Args:
            expr: Expression string
            
        Returns:
            True if expression is safe
        """
        # Only allow basic arithmetic and numbers
        allowed_chars = set('0123456789+-*/()._ ')
        return all(c in allowed_chars for c in expr) and not any(
            dangerous in expr.lower() 
            for dangerous in ['import', 'exec', 'eval', '__', 'open', 'file']
        )
        
            
    def _map_column(self, source_column: str, input_df: pd.DataFrame, col_config: Dict[str, Any] = None) -> List[Any]:
        """
        Map single column directly.
        
        Args:
            source_column: Source column name
            input_df: Input DataFrame
            col_config: Column configuration for transformations
            
        Returns:
            List of mapped values
        """
        if source_column in input_df.columns:
            values = input_df[source_column].tolist()
            
            # Apply transformations if column config is provided
            if col_config:
                formatting = col_config.get("formatting", {})
                
                # Remove asterisks if enabled
                if formatting.get("remove_asterisks", False):
                    values = [str(value).replace("*", "") if pd.notna(value) else value for value in values]
            
            return values
        else:
            self.logger.warning(f"Column '{source_column}' not found in input data")
            return [""] * len(input_df)
            
    def _apply_void_filtering(self, df: pd.DataFrame, config: Dict[str, Any]) -> pd.DataFrame:
        """
        Apply void filtering to remove rows where specified columns are all zero.
        
        Args:
            df: DataFrame to filter (input DataFrame)
            config: Configuration containing void settings
            
        Returns:
            Filtered DataFrame
        """
        try:
            void_config = config.get("void", {})
            
            if not void_config.get("enabled", False):
                return df
                
            zero_columns = void_config.get("zero_columns", [])
            if not zero_columns:
                return df
                
            # Check which columns exist in the input DataFrame
            existing_columns = [col for col in zero_columns if col in df.columns]
            if not existing_columns:
                self.logger.warning(f"No void filter columns found in input data: {zero_columns}")
                self.logger.info(f"Available input columns: {list(df.columns)}")
                return df
                
            self.logger.info(f"Applying void filtering on input columns: {existing_columns}")
            
            # Debug: Show sample values from the columns being checked
            for col in existing_columns:
                sample_values = df[col].head(10).tolist()
                self.logger.info(f"Sample values from '{col}': {sample_values}")
            
            # Create mask for rows where all specified columns are zero
            mask = pd.Series([True] * len(df))
            
            for col in existing_columns:
                numeric_col = pd.to_numeric(df[col], errors='coerce').fillna(0)
                col_mask = (numeric_col == 0)
                self.logger.info(f"Column '{col}': {col_mask.sum()} zero values out of {len(df)} rows")
                mask = mask & col_mask
                
            # Count and remove void rows
            void_rows = mask.sum()
            if void_rows > 0:
                self.logger.info(f"Removing {void_rows} void rows from input data")
                # Debug: Show which rows are being removed
                void_row_indices = df[mask].index.tolist()[:10]  # Show first 10
                self.logger.info(f"Sample void row indices: {void_row_indices}")
                return df[~mask].copy()
            else:
                self.logger.info("No void rows found in input data")
                return df
                
        except Exception as e:
            self.logger.error(f"Error applying void filtering: {str(e)}")
            return df
            
    def process_file(self, input_file: Path, output_dir: Path, config: Dict[str, Any]) -> Path:
        """
        Process a complete file from input to formatted output.
        
        Args:
            input_file: Path to input file
            output_dir: Output directory path
            config: Processing configuration
            
        Returns:
            Path to generated output file
        """
        try:
            self.logger.info(f"Processing file: {input_file.name}")
            
            # Read input file
            input_df = self.read_excel_file(input_file)
            
            # Apply mapping
            output_df = self.apply_mapping(input_df, config)
            
            # Generate output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"{input_file.stem}_formatted_{timestamp}.xlsx"
            output_path = output_dir / output_filename
            
            # Save with formatting
            self.save_formatted_output(output_df, output_path, config)
            
            self.logger.info(f"File processed successfully: {output_filename}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error processing file: {str(e)}")
            raise
            
    def save_formatted_output(self, df: pd.DataFrame, output_path: Path, config: Dict[str, Any]):
        """
        Save DataFrame to Excel with formatting applied.
        
        Args:
            df: DataFrame to save
            output_path: Output file path
            config: Formatting configuration
        """
        try:
            self.logger.info(f"Saving formatted output to: {output_path}")
            
            # Save basic Excel file
            df.to_excel(output_path, index=False, engine='openpyxl')
            
            # Apply formatting
            try:
                self.apply_excel_formatting(output_path, df, config)
            except Exception as e:
                self.logger.warning(f"Excel formatting failed, but file was saved: {str(e)}")
                # Continue without formatting
            
            self.logger.info("Formatted output saved successfully")
            
        except Exception as e:
            self.logger.error(f"Error saving formatted output: {str(e)}")
            raise
            
    def apply_excel_formatting(self, output_path: Path, df: pd.DataFrame, config: Dict[str, Any]):
        """
        Apply Excel formatting to the saved file.
        
        Args:
            output_path: Path to Excel file
            df: DataFrame that was saved
            config: Formatting configuration
        """
        try:
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Apply header formatting
            self._format_headers(ws, df, config)
            
            # Apply column formatting
            self._format_columns(ws, df, config)
            
            # Apply general settings
            self._apply_general_settings(ws, config)
            
            wb.save(output_path)
            self.logger.info("Excel formatting applied successfully")
            
        except Exception as e:
            self.logger.error(f"Error applying Excel formatting: {str(e)}")
            raise
            
    def _format_headers(self, ws, df: pd.DataFrame, config: Dict[str, Any]):
        """Apply header formatting."""
        header_config = config.get("header_formatting", {})
        
        # Header styling
        header_font = Font(
            bold=header_config.get("bold", True),
            color=header_config.get("font_color", "FFFFFF")
        )
        
        header_fill = PatternFill(
            start_color=header_config.get("background_color", "366092"),
            end_color=header_config.get("background_color", "366092"),
            fill_type="solid"
        )
        
        header_alignment = Alignment(
            horizontal=header_config.get("alignment", "center"),
            vertical="center"
        )
        
        # Apply to header row
        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
    def _format_columns(self, ws, df: pd.DataFrame, config: Dict[str, Any]):
        """Apply column-specific formatting."""
        try:
            output_columns = config.get("output_columns", [])
            self.logger.info(f"Formatting {len(output_columns)} columns for DataFrame with {len(df.columns)} columns")
            self.logger.info(f"DataFrame columns: {list(df.columns)}")
            self.logger.info(f"DataFrame shape: {df.shape}")
            
            for col_idx, col_config in enumerate(output_columns, 1):
                self.logger.info(f"Processing column {col_idx}: {col_config.get('name', 'Unknown')}")
                
                if col_idx > len(df.columns):
                    self.logger.warning(f"Column index {col_idx} exceeds DataFrame columns {len(df.columns)}")
                    break
                
                # Column width - auto-fit if enabled, otherwise use configured width
                general_config = config.get("general_settings", {})
                if general_config.get("auto_fit_columns", True):
                    # Auto-fit column width
                    self._auto_fit_column(ws, df, col_idx)
                else:
                    # Use configured width
                    width = col_config.get("width", 15)
                    ws.column_dimensions[get_column_letter(col_idx)].width = width
                
                # Column alignment
                alignment = Alignment(
                    horizontal=col_config.get("alignment", "left"),
                    vertical="center"
                )
                
                # Apply to data rows
                for row_num in range(2, len(df) + 2):
                    try:
                        cell = ws.cell(row=row_num, column=col_idx)
                        cell.alignment = alignment
                        
                        # Apply number formatting
                        formatting = col_config.get("formatting", {})
                        if "number_format" in formatting:
                            cell.number_format = formatting["number_format"]
                    except Exception as cell_error:
                        self.logger.error(f"Error processing cell row {row_num}, col {col_idx}: {str(cell_error)}")
                        raise
                            
        except Exception as e:
            self.logger.error(f"Error in _format_columns: {str(e)}")
            self.logger.error(f"Error type: {type(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            raise
    
    def _auto_fit_column(self, ws, df: pd.DataFrame, col_idx: int):
        """Auto-fit column width based on content."""
        try:
            column_letter = get_column_letter(col_idx)
            
            # Ensure we have valid column index
            if col_idx < 1 or col_idx > len(df.columns):
                self.logger.warning(f"Invalid column index {col_idx}, DataFrame has {len(df.columns)} columns")
                ws.column_dimensions[column_letter].width = 15
                return
                
            col_name = df.columns[col_idx - 1]
            
            # Calculate max width needed
            max_width = len(str(col_name))  # Header width
            
            # Check data rows - limit to first 100 rows for performance
            sample_size = min(len(df), 100)
            for row_idx in range(sample_size):
                try:
                    cell_value = str(df.iloc[row_idx, col_idx - 1])
                    max_width = max(max_width, len(cell_value))
                except (IndexError, KeyError) as e:
                    self.logger.warning(f"Error accessing row {row_idx}, col {col_idx}: {str(e)}")
                    continue
            
            # Add some padding and set reasonable limits
            adjusted_width = min(max(max_width + 2, 8), 50)  # Min 8, Max 50
            ws.column_dimensions[column_letter].width = adjusted_width
            
            self.logger.info(f"Auto-fitted column {col_name} to width {adjusted_width}")
            
        except Exception as e:
            self.logger.warning(f"Error auto-fitting column {col_idx}: {str(e)}")
            # Fallback to default width
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
                    
    def _apply_general_settings(self, ws, config: Dict[str, Any]):
        """Apply general worksheet settings."""
        general_config = config.get("general_settings", {})
        
        # Freeze panes
        freeze_panes = general_config.get("freeze_panes")
        if freeze_panes:
            ws.freeze_panes = freeze_panes